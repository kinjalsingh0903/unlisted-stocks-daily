"""Populate the template with today's top 15 and save as dated file.

Template layout (v6):
  Row 1   = title bar
  Row 2   = metadata strip (Refreshed: auto-fills into B2)
  Row 3   = group band (IDENTITY / VALUATION / ...)
  Row 4   = detailed column headers
  Row 5-19 = 15 data rows
  Row 21  = SUMMARY strip (formulas auto-calc)
"""
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

from src.config import TEMPLATE_PATH

DATA_FIRST_ROW = 5   # first data row
N_PICKS = 15


def populate(picks, fetched_source_names, output_path):
    today = datetime.today()
    wb = load_workbook(TEMPLATE_PATH)

    # === Top 15 sheet ===
    ws = wb["Top 15"]
    ws["B2"] = today.strftime("%d %b %Y")
    ws["B2"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    for i, pick in enumerate(picks[:N_PICKS]):
        row = DATA_FIRST_ROW + i
        ws.cell(row=row, column=2,  value=pick.get("name"))
        ws.cell(row=row, column=3,  value=pick.get("sector"))
        ws.cell(row=row, column=4,  value=pick.get("price"))
        ws.cell(row=row, column=5,  value=pick.get("face_value"))
        ws.cell(row=row, column=6,  value=pick.get("lot_size"))
        # col 7 = Min Investment formula (already in template)
        ws.cell(row=row, column=8,  value=pick.get("market_cap_cr"))
        ws.cell(row=row, column=9,  value=pick.get("w52_high"))
        ws.cell(row=row, column=10, value=pick.get("w52_low"))
        ws.cell(row=row, column=11, value=pick.get("pe"))
        ws.cell(row=row, column=12, value=pick.get("pb"))
        ws.cell(row=row, column=13, value=pick.get("book_value"))
        ws.cell(row=row, column=14, value=pick.get("de"))
        ws.cell(row=row, column=15, value=pick.get("roe_pct"))
        ws.cell(row=row, column=16, value=pick.get("rev_growth_pct"))
        ws.cell(row=row, column=17, value=pick.get("ret_1y_pct"))
        ws.cell(row=row, column=18, value=pick.get("ret_2y_pct"))
        ws.cell(row=row, column=19, value=pick.get("peers_name"))
        ws.cell(row=row, column=20, value=pick.get("peer_pe"))
        ws.cell(row=row, column=21, value=pick.get("ipo_status"))
        ws.cell(row=row, column=22, value=pick.get("ipo_window"))
        ws.cell(row=row, column=23, value=pick.get("n_sources"))
        ws.cell(row=row, column=24, value=pick.get("primary_source"))
        ws.cell(row=row, column=25, value=today)

    # === Sources sheet: stamp Last Fetched (column 7 now, status col is 1) ===
    ws2 = wb["Sources"]
    for row in range(4, 38):
        name = ws2.cell(row=row, column=3).value  # name is now in column 3 (after status + #)
        if name and name in fetched_source_names:
            ws2.cell(row=row, column=7, value=today)

    wb.save(output_path)
    return output_path
